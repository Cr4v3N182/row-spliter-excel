import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox

# Liczenie wierszy w pliku
def get_num_rows_in_xlsx(input_file):
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
    num_rows = sheet.max_row
    wb.close()
    return num_rows

# Dzielenie pliku
def divide_xlsx_file(input_file, max_rows_per_file):
    # Wczytaj plik XLSX
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Podziel arkusz na mniejsze pliki
    num_rows = sheet.max_row
    num_files = (num_rows - 1) // max_rows_per_file + 1

    # Pobierz pierwszy wiersz (nagłówek) z oryginalnego arkusza
    header_row = []
    for cell in sheet[1]:
        header_row.append(cell.value)

    for i in range(num_files):
        start_row = i * max_rows_per_file + 1
        end_row = min((i + 1) * max_rows_per_file, num_rows)

        # Utwórz nowy plik XLSX dla danego zakresu wierszy
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active

        # Dodaj nagłówek do nowego arkusza (oprócz pierwszego pliku)
        if i > 0:
            new_sheet.append(header_row)

        # Skopiuj zawartość z oryginalnego arkusza do nowego
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            new_sheet.append(row)

        # Zapisz plik z odpowiednią nazwą
        base_filename, ext = os.path.splitext(input_file)
        new_filename = f"{base_filename}_{i+1}{ext}"
        new_wb.save(new_filename)
        new_wb.close()

    wb.close()

    messagebox.showinfo("Podział pliku XLSX", "Podział pliku został zakończony.")

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        num_rows = get_num_rows_in_xlsx(file_path)
        info_label.config(text=f"Liczba wierszy w pliku: {num_rows}")
        divide_button.config(state=tk.NORMAL, command=lambda: divide_file_and_disable(file_path))

def divide_file_and_disable(file_path):
    max_rows_per_file = int(max_rows_entry.get())
    divide_xlsx_file(file_path, max_rows_per_file)
    divide_button.config(state=tk.DISABLED)

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Podział pliku XLSX")

    instructions_label = tk.Label(root, text="Wybierz plik XLSX do podzielenia:")
    instructions_label.pack(pady=10)

    browse_button = tk.Button(root, text="Wybierz plik XLSX", command=browse_file)
    browse_button.pack(pady=5)

    info_label = tk.Label(root, text="")
    info_label.pack(pady=5)

    max_rows_label = tk.Label(root, text="Podaj maksymalną liczbę wierszy na plik:")
    max_rows_label.pack(pady=5)

    max_rows_entry = tk.Entry(root)
    max_rows_entry.pack(pady=5)

    divide_button = tk.Button(root, text="Dziel", state=tk.DISABLED)
    divide_button.pack(pady=5)

    root.mainloop()
